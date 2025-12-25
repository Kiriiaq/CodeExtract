"""
Folder Scanner Module - Scan directories and generate structure reports.
Creates comprehensive directory trees with file contents.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Callable
from concurrent.futures import ThreadPoolExecutor


@dataclass
class FileEntry:
    """Information about a file in the scan."""
    path: str
    name: str
    size: int
    modified: datetime
    extension: str
    is_binary: bool = False
    content: Optional[str] = None
    encoding: str = "utf-8"
    error: Optional[str] = None


@dataclass
class DirectoryEntry:
    """Information about a directory in the scan."""
    path: str
    name: str
    files: List[FileEntry] = field(default_factory=list)
    subdirectories: List['DirectoryEntry'] = field(default_factory=list)
    total_files: int = 0
    total_size: int = 0


@dataclass
class ScanResult:
    """Result of a directory scan."""
    root_path: str
    root_entry: Optional[DirectoryEntry] = None
    total_files: int = 0
    total_directories: int = 0
    total_size: int = 0
    scan_time: float = 0.0
    errors: List[str] = field(default_factory=list)


class FolderScanner:
    """
    Scan directories and generate comprehensive reports.
    Supports filtering, content extraction, and multiple output formats.
    """

    DEFAULT_EXCLUDED_DIRS = {
        '__pycache__', '.git', '.svn', '.hg', 'node_modules',
        '.venv', 'venv', 'env', '.idea', '.vscode', 'dist', 'build',
        '.eggs', '*.egg-info', '.tox', '.pytest_cache', '.mypy_cache'
    }

    DEFAULT_EXCLUDED_EXTENSIONS = {
        '.exe', '.dll', '.so', '.dylib', '.o', '.obj',
        '.pyc', '.pyo', '.pyd', '.class',
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.ico', '.svg',
        '.mp3', '.mp4', '.avi', '.mov', '.wav',
        '.zip', '.tar', '.gz', '.rar', '.7z',
        '.pdf', '.doc', '.docx', '.xls', '.xlsx'
    }

    BINARY_EXTENSIONS = {
        '.exe', '.dll', '.so', '.dylib', '.o', '.obj',
        '.pyc', '.pyo', '.pyd', '.class',
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.ico',
        '.mp3', '.mp4', '.avi', '.mov', '.wav',
        '.zip', '.tar', '.gz', '.rar', '.7z', '.pdf'
    }

    def __init__(self):
        self.excluded_dirs: Set[str] = self.DEFAULT_EXCLUDED_DIRS.copy()
        self.excluded_extensions: Set[str] = self.DEFAULT_EXCLUDED_EXTENSIONS.copy()
        self.max_file_size: int = 1024 * 1024  # 1MB
        self.include_content: bool = True
        self.include_binary: bool = False
        self.on_progress: Optional[Callable[[str], None]] = None
        self._should_stop = False

    def configure(self,
                  excluded_dirs: Optional[Set[str]] = None,
                  excluded_extensions: Optional[Set[str]] = None,
                  max_file_size: int = 1024 * 1024,
                  include_content: bool = True,
                  include_binary: bool = False) -> None:
        """Configure scanner options."""
        if excluded_dirs is not None:
            self.excluded_dirs = excluded_dirs
        if excluded_extensions is not None:
            self.excluded_extensions = excluded_extensions
        self.max_file_size = max_file_size
        self.include_content = include_content
        self.include_binary = include_binary

    def stop(self) -> None:
        """Stop the current scan."""
        self._should_stop = True

    def scan(self, directory: str) -> ScanResult:
        """
        Scan a directory and return a complete structure.

        Args:
            directory: Root directory to scan

        Returns:
            ScanResult with complete directory structure
        """
        self._should_stop = False
        start_time = datetime.now()

        result = ScanResult(root_path=directory)

        if not os.path.exists(directory):
            result.errors.append(f"Directory not found: {directory}")
            return result

        if not os.path.isdir(directory):
            result.errors.append(f"Not a directory: {directory}")
            return result

        try:
            result.root_entry = self._scan_directory(directory, result)
            if result.root_entry:
                result.total_files = self._count_files(result.root_entry)
                result.total_directories = self._count_directories(result.root_entry)
                result.total_size = self._calculate_size(result.root_entry)
        except Exception as e:
            result.errors.append(f"Scan error: {str(e)}")

        result.scan_time = (datetime.now() - start_time).total_seconds()
        return result

    def _scan_directory(self, path: str, result: ScanResult) -> Optional[DirectoryEntry]:
        """Recursively scan a directory."""
        if self._should_stop:
            return None

        name = os.path.basename(path) or path

        if self.on_progress:
            self.on_progress(f"Scanning: {path}")

        entry = DirectoryEntry(path=path, name=name)

        try:
            items = sorted(os.listdir(path))
        except PermissionError:
            result.errors.append(f"Permission denied: {path}")
            return entry
        except Exception as e:
            result.errors.append(f"Error reading {path}: {str(e)}")
            return entry

        # Separate directories and files
        dirs = []
        files = []

        for item in items:
            if self._should_stop:
                break

            item_path = os.path.join(path, item)

            if os.path.isdir(item_path):
                if item not in self.excluded_dirs:
                    dirs.append(item)
            elif os.path.isfile(item_path):
                ext = os.path.splitext(item)[1].lower()
                if ext not in self.excluded_extensions:
                    files.append(item)

        # Process subdirectories
        for dir_name in dirs:
            if self._should_stop:
                break
            dir_path = os.path.join(path, dir_name)
            subdir = self._scan_directory(dir_path, result)
            if subdir:
                entry.subdirectories.append(subdir)

        # Process files
        for file_name in files:
            if self._should_stop:
                break
            file_path = os.path.join(path, file_name)
            file_entry = self._scan_file(file_path)
            entry.files.append(file_entry)

        return entry

    def _scan_file(self, path: str) -> FileEntry:
        """Scan a single file."""
        name = os.path.basename(path)
        ext = os.path.splitext(name)[1].lower()

        try:
            stats = os.stat(path)
            size = stats.st_size
            modified = datetime.fromtimestamp(stats.st_mtime)
        except Exception as e:
            return FileEntry(
                path=path,
                name=name,
                size=0,
                modified=datetime.now(),
                extension=ext,
                error=str(e)
            )

        is_binary = ext in self.BINARY_EXTENSIONS

        entry = FileEntry(
            path=path,
            name=name,
            size=size,
            modified=modified,
            extension=ext,
            is_binary=is_binary
        )

        # Read content if requested
        if self.include_content and size <= self.max_file_size:
            if not is_binary or self.include_binary:
                content, encoding, error = self._read_file(path, is_binary)
                entry.content = content
                entry.encoding = encoding
                entry.error = error

        return entry

    def _read_file(self, path: str, is_binary: bool) -> tuple:
        """Read file content with encoding detection."""
        if is_binary:
            try:
                with open(path, 'rb') as f:
                    data = f.read(1024)  # First 1KB only
                hex_dump = ' '.join(f'{b:02x}' for b in data)
                return f"[Binary content - {len(data)} bytes shown]\n{hex_dump}", "binary", None
            except Exception as e:
                return None, "binary", str(e)

        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']

        for encoding in encodings:
            try:
                with open(path, 'r', encoding=encoding) as f:
                    content = f.read()
                return content, encoding, None
            except UnicodeDecodeError:
                continue
            except Exception as e:
                return None, "utf-8", str(e)

        return None, "unknown", "Unable to decode file"

    def _count_files(self, entry: DirectoryEntry) -> int:
        """Count total files in a directory entry."""
        count = len(entry.files)
        for subdir in entry.subdirectories:
            count += self._count_files(subdir)
        return count

    def _count_directories(self, entry: DirectoryEntry) -> int:
        """Count total directories in a directory entry."""
        count = len(entry.subdirectories)
        for subdir in entry.subdirectories:
            count += self._count_directories(subdir)
        return count

    def _calculate_size(self, entry: DirectoryEntry) -> int:
        """Calculate total size of a directory entry."""
        size = sum(f.size for f in entry.files)
        for subdir in entry.subdirectories:
            size += self._calculate_size(subdir)
        return size

    def generate_tree(self, result: ScanResult, include_files: bool = True) -> str:
        """Generate a text tree representation."""
        if not result.root_entry:
            return "No data"

        lines = []
        self._generate_tree_lines(result.root_entry, lines, "", True, include_files)
        return "\n".join(lines)

    def _generate_tree_lines(self, entry: DirectoryEntry, lines: List[str],
                             prefix: str, is_last: bool, include_files: bool) -> None:
        """Generate tree lines recursively."""
        connector = "└── " if is_last else "├── "
        lines.append(f"{prefix}{connector}{entry.name}/")

        new_prefix = prefix + ("    " if is_last else "│   ")

        # Process subdirectories
        items = list(entry.subdirectories)
        if include_files:
            items.extend(entry.files)

        for i, item in enumerate(items):
            is_item_last = (i == len(items) - 1)

            if isinstance(item, DirectoryEntry):
                self._generate_tree_lines(item, lines, new_prefix, is_item_last, include_files)
            elif isinstance(item, FileEntry):
                item_connector = "└── " if is_item_last else "├── "
                size_str = self._format_size(item.size)
                lines.append(f"{new_prefix}{item_connector}{item.name} ({size_str})")

    def _format_size(self, size: int) -> str:
        """Format size to human-readable string."""
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.1f} {unit}" if unit != 'B' else f"{size} {unit}"
            size /= 1024
        return f"{size:.1f} TB"

    def export_to_file(self, result: ScanResult, output_path: str,
                       include_content: bool = True) -> None:
        """Export scan result to a text file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\ufeff')  # BOM for UTF-8
            f.write(f"DIRECTORY SCAN REPORT\n")
            f.write(f"{'=' * 80}\n")
            f.write(f"Root: {result.root_path}\n")
            f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Total Files: {result.total_files}\n")
            f.write(f"Total Directories: {result.total_directories}\n")
            f.write(f"Total Size: {self._format_size(result.total_size)}\n")
            f.write(f"Scan Time: {result.scan_time:.2f}s\n")
            f.write(f"{'=' * 80}\n\n")

            # Directory tree
            f.write("DIRECTORY STRUCTURE:\n")
            f.write("-" * 40 + "\n")
            f.write(self.generate_tree(result))
            f.write("\n\n")

            # File contents
            if include_content and result.root_entry:
                f.write(f"{'=' * 80}\n")
                f.write("FILE CONTENTS:\n")
                f.write(f"{'=' * 80}\n\n")
                self._write_contents(f, result.root_entry)

            # Errors
            if result.errors:
                f.write(f"\n{'=' * 80}\n")
                f.write("ERRORS:\n")
                f.write("-" * 40 + "\n")
                for error in result.errors:
                    f.write(f"  - {error}\n")

    def _write_contents(self, f, entry: DirectoryEntry) -> None:
        """Write file contents to output."""
        for file_entry in entry.files:
            if file_entry.content:
                f.write(f"\n{'#' * 80}\n")
                f.write(f"# FILE: {file_entry.path}\n")
                f.write(f"# Size: {self._format_size(file_entry.size)}\n")
                f.write(f"# Encoding: {file_entry.encoding}\n")
                f.write(f"{'#' * 80}\n\n")
                f.write(file_entry.content)
                f.write("\n\n")

        for subdir in entry.subdirectories:
            self._write_contents(f, subdir)

    def get_all_files_flat(self, entry: DirectoryEntry, root_path: str = "") -> List[Dict]:
        """Get all files as a flat list of dictionaries for export."""
        files = []
        base_path = root_path or entry.path

        for file_entry in entry.files:
            rel_path = os.path.relpath(file_entry.path, base_path)
            files.append({
                'name': file_entry.name,
                'path': file_entry.path,
                'relative_path': rel_path,
                'directory': os.path.dirname(rel_path) or '.',
                'extension': file_entry.extension,
                'size': file_entry.size,
                'size_formatted': self._format_size(file_entry.size),
                'modified': file_entry.modified.strftime('%Y-%m-%d %H:%M:%S'),
                'is_binary': file_entry.is_binary,
                'encoding': file_entry.encoding,
                'has_content': file_entry.content is not None,
                'error': file_entry.error or ''
            })

        for subdir in entry.subdirectories:
            files.extend(self.get_all_files_flat(subdir, base_path))

        return files

    def export_to_excel(self, result: ScanResult, output_path: str) -> bool:
        """
        Export scan result to an Excel file.

        Args:
            result: ScanResult from a scan
            output_path: Path for the Excel file

        Returns:
            True if successful, False otherwise
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fallback to CSV if openpyxl not available
            csv_path = output_path.replace('.xlsx', '.csv').replace('.xls', '.csv')
            return self._export_to_csv_fallback(result, csv_path)

        if not result.root_entry:
            return False

        wb = openpyxl.Workbook()

        # ===== Sheet 1: Summary =====
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary data
        summary_data = [
            ("Property", "Value"),
            ("Root Directory", result.root_path),
            ("Scan Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ("Total Files", result.total_files),
            ("Total Directories", result.total_directories),
            ("Total Size", self._format_size(result.total_size)),
            ("Total Size (bytes)", result.total_size),
            ("Scan Time", f"{result.scan_time:.2f} seconds"),
            ("Errors", len(result.errors))
        ]

        for row_idx, (prop, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=prop)
            ws_summary.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:
                ws_summary.cell(row=row_idx, column=1).font = header_font
                ws_summary.cell(row=row_idx, column=1).fill = header_fill
                ws_summary.cell(row=row_idx, column=2).font = header_font
                ws_summary.cell(row=row_idx, column=2).fill = header_fill

        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 50

        # ===== Sheet 2: Files =====
        ws_files = wb.create_sheet("Files")

        headers = ["Name", "Relative Path", "Directory", "Extension", "Size (bytes)", "Size", "Modified", "Binary", "Encoding"]
        for col, header in enumerate(headers, 1):
            cell = ws_files.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        # Get all files
        all_files = self.get_all_files_flat(result.root_entry)

        for row_idx, file_info in enumerate(all_files, 2):
            ws_files.cell(row=row_idx, column=1, value=file_info['name'])
            ws_files.cell(row=row_idx, column=2, value=file_info['relative_path'])
            ws_files.cell(row=row_idx, column=3, value=file_info['directory'])
            ws_files.cell(row=row_idx, column=4, value=file_info['extension'])
            ws_files.cell(row=row_idx, column=5, value=file_info['size'])
            ws_files.cell(row=row_idx, column=6, value=file_info['size_formatted'])
            ws_files.cell(row=row_idx, column=7, value=file_info['modified'])
            ws_files.cell(row=row_idx, column=8, value="Yes" if file_info['is_binary'] else "No")
            ws_files.cell(row=row_idx, column=9, value=file_info['encoding'])

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            max_length = max(len(str(ws_files.cell(row=r, column=col).value or "")) for r in range(1, min(100, len(all_files) + 2)))
            ws_files.column_dimensions[get_column_letter(col)].width = min(max_length + 2, 50)

        # ===== Sheet 3: By Extension =====
        ws_ext = wb.create_sheet("By Extension")

        # Count by extension
        ext_stats = {}
        for file_info in all_files:
            ext = file_info['extension'] or '(no extension)'
            if ext not in ext_stats:
                ext_stats[ext] = {'count': 0, 'size': 0}
            ext_stats[ext]['count'] += 1
            ext_stats[ext]['size'] += file_info['size']

        ext_headers = ["Extension", "File Count", "Total Size (bytes)", "Total Size"]
        for col, header in enumerate(ext_headers, 1):
            cell = ws_ext.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row_idx, (ext, stats) in enumerate(sorted(ext_stats.items(), key=lambda x: x[1]['size'], reverse=True), 2):
            ws_ext.cell(row=row_idx, column=1, value=ext)
            ws_ext.cell(row=row_idx, column=2, value=stats['count'])
            ws_ext.cell(row=row_idx, column=3, value=stats['size'])
            ws_ext.cell(row=row_idx, column=4, value=self._format_size(stats['size']))

        for col in range(1, 5):
            ws_ext.column_dimensions[get_column_letter(col)].width = 20

        # ===== Sheet 4: Directory Tree =====
        ws_tree = wb.create_sheet("Directory Tree")
        tree_lines = self.generate_tree(result, include_files=False).split('\n')
        for row_idx, line in enumerate(tree_lines, 1):
            ws_tree.cell(row=row_idx, column=1, value=line)
        ws_tree.column_dimensions['A'].width = 80

        # ===== Sheet 5: Errors (if any) =====
        if result.errors:
            ws_errors = wb.create_sheet("Errors")
            ws_errors.cell(row=1, column=1, value="Error").font = header_font
            ws_errors.cell(row=1, column=1).fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
            for row_idx, error in enumerate(result.errors, 2):
                ws_errors.cell(row=row_idx, column=1, value=error)
            ws_errors.column_dimensions['A'].width = 100

        # Save workbook
        wb.save(output_path)
        return True

    def _export_to_csv_fallback(self, result: ScanResult, output_path: str) -> bool:
        """Export to CSV as fallback when openpyxl is not available."""
        import csv

        if not result.root_entry:
            return False

        all_files = self.get_all_files_flat(result.root_entry)

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Relative Path", "Directory", "Extension", "Size (bytes)", "Size", "Modified", "Binary", "Encoding"])
            for file_info in all_files:
                writer.writerow([
                    file_info['name'],
                    file_info['relative_path'],
                    file_info['directory'],
                    file_info['extension'],
                    file_info['size'],
                    file_info['size_formatted'],
                    file_info['modified'],
                    "Yes" if file_info['is_binary'] else "No",
                    file_info['encoding']
                ])

        return True
